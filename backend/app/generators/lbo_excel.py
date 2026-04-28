"""LBO Excel workbook generator — LIVE FORMULAS edition.

Inputs are hardcoded in blue cells (PE/IB convention). All projections,
debt paydown, FCF, exit values, MOIC, and IRR are EXCEL FORMULAS that
reference those inputs. Editing any blue cell triggers full recalc.

Sheets:
  1. Summary       - cross-sheet formulas pulling from each scenario
  2. Base Case     - editable formula-driven model
  3. Bull Case     - editable formula-driven model
  4. Bear Case     - editable formula-driven model
  5. Sensitivity   - 5x5 IRR heatmap (computed via Python LBO engine)
  6. Inputs Ref    - raw diligence pipeline data
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _hdr_fill(): return PatternFill("solid", fgColor="0A2540")
def _row_alt(): return PatternFill("solid", fgColor="F5F7FA")
def _input_fill(): return PatternFill("solid", fgColor="EFF6FF")
def _bull_fill(): return PatternFill("solid", fgColor="DCFCE7")
def _base_fill(): return PatternFill("solid", fgColor="FEF3C7")
def _bear_fill(): return PatternFill("solid", fgColor="FEE2E2")
def _white_bold(): return Font(bold=True, color="FFFFFF")
def _bold(): return Font(bold=True)
def _input_font(): return Font(color="1D4ED8", bold=True)
def _border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_scenario_sheet(ws, scenario_name, scenario, fill, financial, market):
    """Build a fully formula-driven LBO scenario sheet.

    Cell layout (memorize):
      B7=StartRev   B8=StartMargin   B9=Cash       B10=ExDebt   B11=MktCap
      B12=Premium%  B13=Leverage%    B14=Growth%   B15=Y5Mgn%   B16=ExitMult
      B17=IntRate%  B18=Capex%       B19=Tax%      B20=Hold

      B23=PurchasePrice   B24=DebtAtClose   B25=SponsorEquity

      Row 29-36 cols B-F: Y1-Y5 projections
      Row 39-44 col B:    Exit & Returns
    """
    inputs = scenario.get("inputs", {}) or {}

    # === Title bar
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{scenario_name.upper()} CASE  -  LIVE LBO MODEL"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = _hdr_fill()
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # === Color legend
    ws["A3"] = "BLUE = hardcoded input (editable)"
    ws["A3"].font = _input_font()
    ws["A4"] = "BLACK = formula (derived from inputs)"
    ws["A4"].font = Font(color="000000", italic=True)

    # === INPUTS section header
    ws["A6"] = "INPUTS  -  EDIT BLUE CELLS TO FLEX ASSUMPTIONS"
    ws["A6"].font = _white_bold()
    ws["A6"].fill = _hdr_fill()
    ws.merge_cells("A6:B6")
    ws["A6"].alignment = Alignment(horizontal="center")

    # Pull diligence data
    rev_y0 = float(financial.get("revenue_usd_millions") or 0)
    margin_y0 = float(financial.get("ebitda_margin_pct") or 30.0)
    cash = float(financial.get("cash_and_equivalents_usd_millions") or 0)
    existing_debt = float(financial.get("total_debt_usd_millions") or 0)
    mkt_cap = float(market.get("market_cap_usd_millions") or 0)
    purchase_price = float(inputs.get("purchase_price_usd_millions") or 0)
    premium_pct = round(((purchase_price / mkt_cap) - 1) * 100, 1) if mkt_cap > 0 else 30.0
    leverage = float(inputs.get("debt_pct") or 55.0)
    growth = float(inputs.get("revenue_growth_pct") or 4.0)
    margin_y5 = float(inputs.get("ebitda_margin_y5_pct") or margin_y0 + 1)
    exit_mult = float(inputs.get("exit_multiple") or 12.0)
    hold = int(inputs.get("hold_period_years") or 5)

    # === Hardcoded inputs (rows 7-20, column B)
    input_rows = [
        ("Starting Revenue ($M)",      rev_y0,        "#,##0"),
        ("Starting EBITDA Margin (%)", margin_y0,     "0.00"),
        ("Cash & Equivalents ($M)",    cash,          "#,##0"),
        ("Existing Debt ($M)",         existing_debt, "#,##0"),
        ("Market Cap ($M)",            mkt_cap,       "#,##0"),
        ("Purchase Premium (%)",       premium_pct,   "0.0"),
        ("Leverage at Close (%)",      leverage,      "0.0"),
        ("Revenue Growth (%)",         growth,        "0.00"),
        ("Y5 EBITDA Margin (%)",       margin_y5,     "0.00"),
        ("Exit Multiple (x)",          exit_mult,     "0.0"),
        ("Interest Rate on Debt (%)",  7.0,           "0.0"),
        ("Capex (% of Revenue)",       3.0,           "0.0"),
        ("Tax Rate (%)",               25.0,          "0.0"),
        ("Hold Period (years)",        hold,          "0"),
    ]
    for i, (k, v, fmt) in enumerate(input_rows, start=7):
        ws.cell(row=i, column=1, value=k).font = _bold()
        ws.cell(row=i, column=1).border = _border()
        c = ws.cell(row=i, column=2, value=v)
        c.font = _input_font()
        c.fill = _input_fill()
        c.number_format = fmt
        c.border = _border()

    # === DERIVED INPUTS header (row 22)
    ws["A22"] = "DERIVED INPUTS (formulas)"
    ws["A22"].font = _white_bold()
    ws["A22"].fill = _hdr_fill()
    ws.merge_cells("A22:B22")
    ws["A22"].alignment = Alignment(horizontal="center")

    # Rows 23-25: derived
    derived = [
        ("Purchase Price ($M)",          "=B11*(1+B12/100)", "#,##0"),
        ("Debt at Close ($M)",           "=B23*B13/100",     "#,##0"),
        ("Sponsor Equity at Close ($M)", "=B23-B24",         "#,##0"),
    ]
    for i, (k, formula, fmt) in enumerate(derived, start=23):
        ws.cell(row=i, column=1, value=k).font = _bold()
        ws.cell(row=i, column=1).border = _border()
        c = ws.cell(row=i, column=2, value=formula)
        c.number_format = fmt
        c.border = _border()
        c.font = _bold()

    # === PROJECTIONS section (rows 27-36)
    ws["A27"] = "5-YEAR PROJECTIONS  -  ALL FORMULAS"
    ws["A27"].font = _white_bold()
    ws["A27"].fill = _hdr_fill()
    ws.merge_cells("A27:F27")
    ws["A27"].alignment = Alignment(horizontal="center")

    # Year headers (row 28)
    ws.cell(row=28, column=1, value="Metric ($M)").font = _white_bold()
    ws.cell(row=28, column=1).fill = _hdr_fill()
    ws.cell(row=28, column=1).border = _border()
    for ci, yr in enumerate(["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"], start=2):
        c = ws.cell(row=28, column=ci, value=yr)
        c.font = _white_bold()
        c.fill = _hdr_fill()
        c.alignment = Alignment(horizontal="center")
        c.border = _border()

    # Row 29: Revenue (compounding growth)
    ws.cell(row=29, column=1, value="Revenue").font = _bold()
    ws.cell(row=29, column=1).border = _border()
    c = ws.cell(row=29, column=2, value="=$B$7*(1+$B$14/100)")
    c.number_format = "#,##0"
    c.border = _border()
    for ci, prev_col in [(3, "B"), (4, "C"), (5, "D"), (6, "E")]:
        c = ws.cell(row=29, column=ci, value=f"={prev_col}29*(1+$B$14/100)")
        c.number_format = "#,##0"
        c.border = _border()

    # Row 30: EBITDA Margin (linear interpolation Y0->Y5)
    ws.cell(row=30, column=1, value="EBITDA Margin (%)").font = _bold()
    ws.cell(row=30, column=1).border = _border()
    for ci, n in [(2, 1), (3, 2), (4, 3), (5, 4), (6, 5)]:
        c = ws.cell(row=30, column=ci, value=f"=$B$8+($B$15-$B$8)*{n}/5")
        c.number_format = "0.00"
        c.border = _border()

    # Row 31: EBITDA = Revenue * Margin / 100
    ws.cell(row=31, column=1, value="EBITDA").font = _bold()
    ws.cell(row=31, column=1).border = _border()
    for ci in range(2, 7):
        col = get_column_letter(ci)
        c = ws.cell(row=31, column=ci, value=f"={col}29*{col}30/100")
        c.number_format = "#,##0"
        c.border = _border()

    # Row 32: Interest Expense (Y1 uses opening debt B24; Y2-Y5 use prior EoY debt B36..E36)
    ws.cell(row=32, column=1, value="Interest Expense").font = _bold()
    ws.cell(row=32, column=1).border = _border()
    c = ws.cell(row=32, column=2, value="=$B$24*$B$17/100")
    c.number_format = "#,##0"
    c.border = _border()
    for ci, prev_col in [(3, "B"), (4, "C"), (5, "D"), (6, "E")]:
        c = ws.cell(row=32, column=ci, value=f"={prev_col}36*$B$17/100")
        c.number_format = "#,##0"
        c.border = _border()

    # Row 33: Capex = Revenue * Capex%
    ws.cell(row=33, column=1, value="Capex").font = _bold()
    ws.cell(row=33, column=1).border = _border()
    for ci in range(2, 7):
        col = get_column_letter(ci)
        c = ws.cell(row=33, column=ci, value=f"={col}29*$B$18/100")
        c.number_format = "#,##0"
        c.border = _border()

    # Row 34: Taxes = MAX(0, (EBITDA - Interest) * Tax%)
    ws.cell(row=34, column=1, value="Taxes").font = _bold()
    ws.cell(row=34, column=1).border = _border()
    for ci in range(2, 7):
        col = get_column_letter(ci)
        c = ws.cell(row=34, column=ci, value=f"=MAX(0,({col}31-{col}32)*$B$19/100)")
        c.number_format = "#,##0"
        c.border = _border()

    # Row 35: Free Cash Flow
    ws.cell(row=35, column=1, value="Free Cash Flow").font = _bold()
    ws.cell(row=35, column=1).border = _border()
    for ci in range(2, 7):
        col = get_column_letter(ci)
        c = ws.cell(row=35, column=ci, value=f"={col}31-{col}32-{col}33-{col}34")
        c.number_format = "#,##0"
        c.border = _border()

    # Row 36: Debt Balance EoY
    ws.cell(row=36, column=1, value="Debt Balance (EoY)").font = _bold()
    ws.cell(row=36, column=1).border = _border()
    c = ws.cell(row=36, column=2, value="=MAX(0,$B$24-MAX(0,B35))")
    c.number_format = "#,##0"
    c.border = _border()
    for ci, prev_col in [(3, "B"), (4, "C"), (5, "D"), (6, "E")]:
        col = get_column_letter(ci)
        c = ws.cell(row=36, column=ci, value=f"=MAX(0,{prev_col}36-MAX(0,{col}35))")
        c.number_format = "#,##0"
        c.border = _border()

    # === EXIT & RETURNS (rows 38-44)
    ws["A38"] = "EXIT & RETURNS (formulas)"
    ws["A38"].font = _white_bold()
    ws["A38"].fill = _hdr_fill()
    ws.merge_cells("A38:B38")
    ws["A38"].alignment = Alignment(horizontal="center")

    exit_rows = [
        ("Exit EBITDA ($M)",        "=F31",                                         "#,##0"),
        ("Exit EV ($M)",            "=B39*$B$16",                                   "#,##0"),
        ("Net Debt at Exit ($M)",   "=F36",                                         "#,##0"),
        ("Exit Equity Value ($M)",  "=MAX(0,B40-B41)",                              "#,##0"),
        ("MOIC",                    "=IFERROR(B42/$B$25,0)",                        '0.00"x"'),
        ("IRR (%)",                 "=IFERROR(((B42/$B$25)^(1/$B$20)-1)*100,0)",    '0.0"%"'),
    ]
    for i, (k, formula, fmt) in enumerate(exit_rows, start=39):
        ws.cell(row=i, column=1, value=k).font = _bold()
        ws.cell(row=i, column=1).border = _border()
        c = ws.cell(row=i, column=2, value=formula)
        c.number_format = fmt
        c.fill = fill
        c.font = Font(bold=True, color="0A2540")
        c.border = _border()

    # Footnote
    ws["A46"] = "Tip: Edit any BLUE input cell. Projections, debt paydown, and returns recalculate live."
    ws["A46"].font = Font(italic=True, color="64748B", size=10)
    ws.merge_cells("A46:F46")

    # Column widths
    ws.column_dimensions["A"].width = 32
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14


def _write_sensitivity_sheet(ws, base_scenario):
    """Sensitivity heatmap: IRR by Exit Multiple x Revenue Growth (Python-computed)."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "SENSITIVITY: BASE-CASE IRR (% by Exit Multiple x Revenue Growth)"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = _hdr_fill()
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    inputs = base_scenario.get("inputs", {})
    exit_mults = [10.0, 11.0, 12.0, 13.0, 14.0]
    growths = [1.0, 3.0, 4.0, 5.0, 7.0]

    ws.cell(row=3, column=1, value="").fill = _hdr_fill()
    for ci, m in enumerate(exit_mults, start=2):
        c = ws.cell(row=3, column=ci, value=f"{m:.1f}x")
        c.font = _white_bold()
        c.fill = _hdr_fill()
        c.alignment = Alignment(horizontal="center")
        c.border = _border()

    from app.services.lbo_engine import run_lbo_scenario

    for ri, growth in enumerate(growths, start=4):
        c = ws.cell(row=ri, column=1, value=f"Growth {growth:.1f}%")
        c.fill = _hdr_fill()
        c.font = _white_bold()
        c.border = _border()

        for ci, m in enumerate(exit_mults, start=2):
            try:
                rev_y0 = base_scenario["projections"]["revenue_usd_millions"][0] / (1 + inputs["revenue_growth_pct"] / 100)
                ebitda_y0 = base_scenario["projections"]["ebitda_usd_millions"][0]
                margin_y0_pct = (ebitda_y0 / rev_y0 * 100.0) if rev_y0 > 0 else 30.0
                debt = inputs["debt_at_close_usd_millions"]
                eq = inputs["sponsor_equity_at_close_usd_millions"]
                tp_total = debt + eq
                result = run_lbo_scenario(
                    ebitda_y0=ebitda_y0,
                    revenue_y0=rev_y0,
                    take_private_price_total=tp_total,
                    cash_on_balance_sheet=0,
                    existing_debt=0,
                    revenue_growth_pct=growth,
                    ebitda_margin_y5_pct=margin_y0_pct + 1.0,
                    exit_multiple=m,
                    debt_pct_of_purchase=eq / (debt + eq) if (debt + eq) > 0 else 0.55,
                )
                irr = result["returns"]["irr_pct"]
            except Exception:
                irr = None
            cc = ws.cell(row=ri, column=ci, value=irr)
            cc.number_format = '0.0"%"'
            cc.alignment = Alignment(horizontal="center")
            cc.border = _border()
            if isinstance(irr, (int, float)):
                if irr >= 25: cc.fill = PatternFill("solid", fgColor="86EFAC")
                elif irr >= 20: cc.fill = PatternFill("solid", fgColor="BBF7D0")
                elif irr >= 15: cc.fill = PatternFill("solid", fgColor="FEF3C7")
                elif irr >= 10: cc.fill = PatternFill("solid", fgColor="FED7AA")
                else: cc.fill = PatternFill("solid", fgColor="FECACA")

    ws.column_dimensions["A"].width = 18
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 12


def generate_lbo_excel(
    ticker: str,
    company_name: str,
    output_dir: str,
    lbo: dict,
    financial: dict,
    market: dict,
) -> str:
    os.makedirs(output_dir, exist_ok=True)
    filename = f"{ticker}_LBO_Model_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    wb.remove(wb.active)

    # === Summary tab (cross-sheet formulas)
    ws_sum = wb.create_sheet("Summary")
    ws_sum.merge_cells("A1:E1")
    ws_sum["A1"] = f"{ticker}  -  {company_name}: LBO Returns Summary"
    ws_sum["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_sum["A1"].fill = _hdr_fill()
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 28

    ws_sum["A2"] = "Live formulas reference each scenario sheet — edit inputs there to flex."
    ws_sum["A2"].font = Font(italic=True, color="64748B", size=10)
    ws_sum.merge_cells("A2:E2")

    headers = ["Scenario", "IRR (%)", "MOIC", "Exit Equity ($M)", "Exit EV ($M)"]
    for ci, h in enumerate(headers, start=1):
        c = ws_sum.cell(row=4, column=ci, value=h)
        c.font = _white_bold()
        c.fill = _hdr_fill()
        c.border = _border()
        c.alignment = Alignment(horizontal="center")

    if lbo.get("status") == "ok":
        scenario_rows = [
            ("Base", "Base Case", _base_fill()),
            ("Bull", "Bull Case", _bull_fill()),
            ("Bear", "Bear Case", _bear_fill()),
        ]
        for i, (name, sheet, fill) in enumerate(scenario_rows, start=5):
            ws_sum.cell(row=i, column=1, value=name).fill = fill
            # Cross-sheet formulas
            c_irr = ws_sum.cell(row=i, column=2, value=f"='{sheet}'!B44")
            c_moic = ws_sum.cell(row=i, column=3, value=f"='{sheet}'!B43")
            c_eq = ws_sum.cell(row=i, column=4, value=f"='{sheet}'!B42")
            c_ev = ws_sum.cell(row=i, column=5, value=f"='{sheet}'!B40")
            for c in (c_irr, c_moic, c_eq, c_ev):
                c.fill = fill
                c.border = _border()
            c_irr.number_format = '0.0"%"'
            c_moic.number_format = '0.00"x"'
            c_eq.number_format = "#,##0"
            c_ev.number_format = "#,##0"
            ws_sum.cell(row=i, column=1).border = _border()

    for col in range(1, 6):
        ws_sum.column_dimensions[get_column_letter(col)].width = 18

    # === Per-scenario tabs (formula-driven)
    if lbo.get("status") == "ok":
        ws_base = wb.create_sheet("Base Case")
        _write_scenario_sheet(ws_base, "Base", lbo["base"], _base_fill(), financial, market)

        ws_bull = wb.create_sheet("Bull Case")
        _write_scenario_sheet(ws_bull, "Bull", lbo["bull"], _bull_fill(), financial, market)

        ws_bear = wb.create_sheet("Bear Case")
        _write_scenario_sheet(ws_bear, "Bear", lbo["bear"], _bear_fill(), financial, market)

        ws_sens = wb.create_sheet("Sensitivity")
        _write_sensitivity_sheet(ws_sens, lbo["base"])

    # === Inputs Reference (raw data, unchanged)
    ws_inputs = wb.create_sheet("Inputs Reference")
    ws_inputs["A1"] = "Inputs from Diligence Pipeline"
    ws_inputs["A1"].font = _bold()
    refs = [
        ("Ticker", ticker),
        ("Company", company_name),
        ("Revenue ($M)", financial.get("revenue_usd_millions")),
        ("EBITDA ($M)", financial.get("ebitda_usd_millions")),
        ("EBITDA Margin (%)", financial.get("ebitda_margin_pct")),
        ("Free Cash Flow ($M)", financial.get("free_cash_flow_usd_millions")),
        ("Total Debt ($M)", financial.get("total_debt_usd_millions")),
        ("Cash ($M)", financial.get("cash_and_equivalents_usd_millions")),
        ("Market Cap ($M)", market.get("market_cap_usd_millions")),
        ("Current Price", market.get("current_price")),
        ("P/E", market.get("pe_ratio")),
        ("EV/EBITDA", market.get("ev_to_ebitda")),
        ("Beta", market.get("beta")),
    ]
    for i, (k, v) in enumerate(refs, start=3):
        ws_inputs.cell(row=i, column=1, value=k).font = _bold()
        ws_inputs.cell(row=i, column=2, value=v)
    ws_inputs.column_dimensions["A"].width = 24
    ws_inputs.column_dimensions["B"].width = 24

    wb.save(filepath)
    return filepath
